import pandas as pd
from pathlib import Path
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows

class ReportGenerator:
    def generate_full_report(self, results: list, job_profile: dict):
        """Generate professional Excel report with charts"""
        output_dir = Path("static/results")
        output_dir.mkdir(exist_ok=True, parents=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_dir / f"CV_Analysis_Report_{timestamp}.xlsx"
        
        # Create professional Excel workbook
        wb = Workbook()
        ws_overview = wb.active
        ws_overview.title = "CV Ranking"
        ws_details = wb.create_sheet("Candidate Details")
        ws_charts = wb.create_sheet("Charts")
        
        # ===== SHEET 1: RANKING TABLE =====
        self._create_ranking_sheet(ws_overview, results, job_profile)
        
        # ===== SHEET 2: DETAILS =====
        self._create_details_sheet(ws_details, results)
        
        # ===== SHEET 3: CHARTS =====
        self._create_charts_sheet(ws_charts, results)
        
        # Save
        wb.save(excel_file)
        
        print(f"📊 Professional Excel saved: {excel_file}")
        return str(excel_file)
    
    def _create_ranking_sheet(self, ws, results, job_profile):
        """Professional ranking table"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Data styling
        data_font = Font(size=11)
        rank_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"AI CV Analysis Report - {job_profile['position']}"
        ws['A1'].font = Font(bold=True, size=20, color="366092")
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['Rank', 'Candidate', 'Score', 'Experience', 'Education', 'Skills Match', 'Languages', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for i, result in enumerate(results, 1):
            row = 5 + i
            score_color = self._get_score_color(result['final_score'])
            
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=result['filename'][:25])
            ws.cell(row=row, column=3, value=f"{result['final_score']:.1f}/100").font = Font(bold=True, color=score_color)
            ws.cell(row=row, column=4, value=f"{result['features'].years_experience:.1f} years")
            ws.cell(row=row, column=5, value=result['features'].education_level.upper())
            ws.cell(row=row, column=6, value=f"{len(result['features'].technical_skills)} skills")
            ws.cell(row=row, column=7, value=', '.join(result['features'].languages[:2]))
            
            # Status
            status = "TOP" if result['final_score'] >= 85 else "SHORTLIST" if result['final_score'] >= 70 else "REVIEW"
            status_color = "4CAF50" if status == "TOP" else "FF9800" if status == "SHORTLIST" else "9E9E9E"
            ws.cell(row=row, column=8, value=status).font = Font(bold=True, color=status_color)
            
            # Rank 1-3 highlighting
            if i <= 3:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = rank_fill
    
    def _create_details_sheet(self, ws, results):
        """Detailed candidate information"""
        headers = ['Filename', 'Total Score', 'Experience Score', 'Skills Score', 'Education', 'Skills Found', 'Languages', 'Companies']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        for i, result in enumerate(results):
            row = 2 + i
            ws.cell(row=row, column=1, value=result['filename'])
            ws.cell(row=row, column=2, value=result['final_score'])
            breakdown = result['breakdown']
            ws.cell(row=row, column=3, value=f"{breakdown['experience']*100:.1f}")
            ws.cell(row=row, column=4, value=f"{breakdown['skills']*100:.1f}")
            ws.cell(row=row, column=5, value=result['features'].education_level)
            ws.cell(row=row, column=6, value=', '.join(result['features'].technical_skills))
            ws.cell(row=row, column=7, value=', '.join(result['features'].languages))
            ws.cell(row=row, column=8, value=result['features'].companies_count)
    
    def _create_charts_sheet(self, ws, results):
        """Charts and analytics"""
        # Prepare data
        df = pd.DataFrame([{
            'Candidate': f"CV_{i+1}",
            'Score': result['final_score'],
            'Experience': result['features'].years_experience,
            'Skills_Count': len(result['features'].technical_skills)
        } for i, result in enumerate(results[:20])])
        
        # Write data to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Top CV Scores"
        chart.y_axis.title = 'Score'
        chart.x_axis.title = 'Candidates'
        
        data = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 15
        chart.height = 10
        ws.add_chart(chart, "D2")
        
        # Score distribution chart
        ws.cell(row=25, column=1, value="Score Distribution").font = Font(bold=True, size=14)
    
    def _get_score_color(self, score):
        """Color coding for scores"""
        if score >= 85:
            return "4CAF50"  # Green
        elif score >= 70:
            return "FF9800"  # Orange
        else:
            return "F44336"  # Red